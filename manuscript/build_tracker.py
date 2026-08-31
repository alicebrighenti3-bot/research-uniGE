import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
EXAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------
# SHEET 1: Legend / instructions
# ---------------------------------------------------------------
ws_leg = wb.active
ws_leg.title = "Legend"
ws_leg.sheet_view.showGridLines = False
ws_leg.column_dimensions["A"].width = 100

title = ws_leg.cell(row=1, column=1, value="APS + Cancer Anticoagulation — Literature Extraction Tracker")
title.font = Font(name=FONT_NAME, bold=True, size=14)

lines = [
    "",
    "PURPOSE",
    "Log every article you (or I) find — RCTs, cohorts, registries, case series, individual case "
    "reports — one ROW PER PATIENT on the 'Data' tab. The 'Dashboard' tab recalculates automatically "
    "as rows are added, so at any point we can see, with real numbers, whether a systematic review "
    "with meta-analysis is defensible, or whether only a descriptive/narrative synthesis is honest.",
    "",
    "HOW TO USE",
    "1. For each article, create one row per patient described (a case report = 1 row; a case series "
    "or cohort of N patients = N rows, sharing the same Study_ID).",
    "2. Cells shaded YELLOW on the 'Data' tab are the ones to fill in. Use the dropdown lists where "
    "provided (Design, Population_group, aPL_profile, Anticoag_strategy, and all Yes/No/NA columns) — "
    "typing free text into those columns breaks the automatic counts on the Dashboard.",
    "3. Rows 2-4 on the 'Data' tab are WORKED EXAMPLES (shaded orange, Is_Example = Y). They are "
    "excluded from every Dashboard count automatically. Leave them as a formatting reference, or "
    "delete them once you're comfortable with the layout — either is fine.",
    "4. The single most important column is 'Population_group' — this is the A / B / C distinction "
    "from the manuscript (Section 1.4): A = simple/transient aPL positivity in an oncology patient, "
    "B = clinically defined APS + cancer, C = high-risk or triple-positive APS + cancer. Everything "
    "the Dashboard tells you about feasibility hinges on getting this column right for each patient, "
    "so when an article is ambiguous, use the Notes column to say why and flag it for a second look.",
    "5. Send me the updated file at any point (or keep adding to it here) and I will re-check the "
    "numbers, sanity-check the Population_group assignments against what the paper actually reports, "
    "and give you an updated verdict on review type and meta-analysis feasibility.",
    "",
    "WHAT THE DASHBOARD CAN AND CANNOT TELL YOU",
    "The thresholds on the Dashboard tab are rule-of-thumb gates, not statistical guarantees. Clearing "
    "them means a pooled analysis becomes worth attempting in principle; it does NOT by itself mean "
    "pooling is valid — that still requires the populations, exposures, and outcomes across studies "
    "to be clinically comparable, which only a full read of each source (not just the count) can "
    "confirm. In particular: case reports and case series have no denominator and no comparator arm, "
    "so accumulating many of them supports at most a DESCRIPTIVE pooled proportion (e.g. 'x/y reported "
    "patients on a VKA had a recurrence') presented with an explicit publication-bias caveat — it does "
    "not support a comparative effect estimate (e.g. a pooled risk ratio of DOAC vs VKA) unless "
    "genuine comparative studies (cohorts, registries, RCTs) with both an exposed and a reference group "
    "are identified.",
]
r = 2
for line in lines:
    if line and line.isupper():
        c = ws_leg.cell(row=r, column=1, value=line)
        c.font = Font(name=FONT_NAME, bold=True, size=11)
    else:
        c = ws_leg.cell(row=r, column=1, value=line)
        c.font = Font(name=FONT_NAME, size=10)
    c.alignment = WRAP
    ws_leg.row_dimensions[r].height = 14 if not line else (30 if len(line) > 90 else 16)
    r += 1

ws_leg.cell(row=r + 1, column=1, value="Colour key").font = Font(name=FONT_NAME, bold=True, size=11)
r += 2
key_rows = [
    ("Yellow fill", INPUT_FILL, "Cell(s) to fill in on the Data tab"),
    ("Orange fill", EXAMPLE_FILL, "Worked example row — excluded from all Dashboard counts"),
]
for label, fill, desc in key_rows:
    c = ws_leg.cell(row=r, column=1, value=f"{label} — {desc}")
    c.fill = fill
    c.font = Font(name=FONT_NAME, size=10)
    r += 1

# ---------------------------------------------------------------
# SHEET 2: Data
# ---------------------------------------------------------------
ws = wb.create_sheet("Data")
ws.sheet_view.showGridLines = True

headers = [
    ("Is_Example", "Y for the worked-example rows only; leave blank for real data."),
    ("Study_ID", "Short key shared by all patients from the same article, e.g. Smith_2022 or Cureus_2024_CaseRep."),
    ("Patient_ID", "Identifier within the study, e.g. 1, 2, '3 of 5'."),
    ("Design", "Case report / Case series / Retrospective cohort / Prospective cohort / Registry / RCT."),
    ("Year", "Publication year."),
    ("Country", "Country of the reporting centre, if stated."),
    ("Age", "Patient age in years at the index event."),
    ("Sex", "M / F / Not stated."),
    ("APS_classification_criteria", "Sapporo 1999 / Sydney 2006 / ACR-EULAR 2023 / Not stated / Not applicable (population A)."),
    ("aPL_profile", "Single / Double / Triple / Not specified."),
    ("Population_group", "A = simple/transient aPL positivity; B = defined APS + cancer; C = high-risk/triple-positive APS + cancer. See Legend."),
    ("Malignancy_type", "Tumour type / haematological malignancy."),
    ("Malignancy_active_treatment", "Chemotherapy / Radiotherapy / Chemo+Radio / Surgery only / None active / Remission."),
    ("Timing_APS_vs_Cancer", "APS diagnosed before cancer / after cancer / concurrently / unclear."),
    ("Index_event_type", "Venous / Arterial / Both / Microvascular."),
    ("Index_event_site", "Free text, e.g. PE, DVT lower limb, stroke, PICC-line thrombosis."),
    ("Anticoag_strategy", "VKA / DOAC / LMWH / UFH / Switched / Other."),
    ("Switch_occurred", "Y / N — was there a documented switch between anticoagulant classes?"),
    ("Thrombotic_recurrence", "Y / N / NA — recurrent VTE or arterial event on treatment."),
    ("Major_bleeding", "Y / N / NA — ISTH major bleeding."),
    ("CRNMB", "Y / N / NA — clinically relevant non-major bleeding."),
    ("Follow_up_months", "Reported follow-up duration in months."),
    ("Mortality", "Y / N / Unclear."),
    ("Full_citation", "Author, title, journal, year — as complete as available."),
    ("Notes", "Anything ambiguous, especially Population_group judgement calls."),
]

for col_idx, (name, desc) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = BORDER
    cell.comment = Comment(desc, "Tracker")
    letter = get_column_letter(col_idx)
    ws.column_dimensions[letter].width = 15 if name not in ("Full_citation", "Notes", "Index_event_site") else 34
ws.row_dimensions[1].height = 30
ws.freeze_panes = "C2"

N_ROWS = 1000  # working range for formulas / validation

# Data validation dropdowns
dv_defs = {
    "A": '"Y"',
    "D": '"Case report,Case series,Retrospective cohort,Prospective cohort,Registry,RCT"',
    "H": '"M,F,Not stated"',
    "J": '"Single,Double,Triple,Not specified"',
    "K": '"A,B,C"',
    "M": '"Chemotherapy,Radiotherapy,Chemo+Radio,Surgery only,None active,Remission"',
    "O": '"Venous,Arterial,Both,Microvascular"',
    "P": '"VKA,DOAC,LMWH,UFH,Switched,Other"',
    "Q": '"Y,N"',
    "S": '"Y,N,NA"',
    "T": '"Y,N,NA"',
    "V": '"Y,N,Unclear"',
}
# R (Thrombotic_recurrence) uses same Y/N/NA list as S/T
dv_defs["R"] = '"Y,N,NA"'

for col_letter, formula in dv_defs.items():
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{N_ROWS}")

# Shade the fillable region yellow for real-data rows (from row 5 onward, after the 3 examples)
for row in range(2, N_ROWS + 1):
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=row, column=col_idx)
        c.border = BORDER
        c.font = Font(name=FONT_NAME, size=10)
        if row >= 5:
            c.fill = INPUT_FILL

# Worked example rows (2-4), shaded orange, Is_Example = Y
examples = [
    ["Y", "Cureus_2024_PerfectStorm", "1", "Case report", 2024, "USA", 51, "M",
     "Not stated", "Not specified", "B", "Duodenal cancer (history of)", "None active",
     "After cancer", "Both", "Ischaemic stroke, ICH, PE", "LMWH", "Y", "N", "N", "N", 3,
     "N", "Cancer History, Antiphospholipid Syndrome, and Lupus Anticoagulant: A Perfect Storm for Thrombosis. Cureus. 2024/2025.",
     "Anticoagulation initially withheld for ICH, then started after PE; single LAC positivity reported, not triple; illustrates the dilemma but is not generalisable."],
    ["Y", "Example_TripleC_Illustrative", "1", "Case report", 2020, "-", 45, "F",
     "Sydney 2006", "Triple", "C", "Breast cancer", "Chemotherapy",
     "Before cancer", "Venous", "Recurrent DVT", "VKA", "N", "N", "N", "N", 6,
     "N", "PLACEHOLDER — replace with a real triple-positive+cancer case once found; format only.",
     "This row shows the exact shape a genuine Population C case needs to take."],
    ["Y", "Example_TripleC_Illustrative", "2", "Case series", 2020, "-", 60, "M",
     "ACR-EULAR 2023", "Triple", "C", "Lung cancer", "Radiotherapy",
     "Concurrently", "Arterial", "Ischaemic stroke", "DOAC", "Y", "Y", "N", "N", 4,
     "Unclear", "PLACEHOLDER — replace with a real triple-positive+cancer case once found; format only.",
     "Switched from DOAC after breakthrough stroke; shows how Switch_occurred/Thrombotic_recurrence pair up."],
]
for i, row_vals in enumerate(examples, start=2):
    for col_idx, val in enumerate(row_vals, start=1):
        c = ws.cell(row=i, column=col_idx, value=val)
        c.fill = EXAMPLE_FILL
        c.font = Font(name=FONT_NAME, size=10, italic=True)
        c.border = BORDER
        c.alignment = WRAP

# Helper column Y: Study_ID key only for real (non-example) rows with a Study_ID filled in
ws.cell(row=1, column=25, value="StudyKey_ifReal (helper — do not edit)").font = HEADER_FONT
ws.cell(row=1, column=25).fill = HEADER_FILL
ws.column_dimensions["Y"].width = 20
for row in range(2, N_ROWS + 1):
    ws.cell(row=row, column=25, value=f'=IF(AND(A{row}<>"Y",B{row}<>""),B{row},"")')

# Same helper restricted to Population_group = C, for the "distinct studies with a C patient" count
ws.cell(row=1, column=26, value="StudyKey_ifRealC (helper — do not edit)").font = HEADER_FONT
ws.cell(row=1, column=26).fill = HEADER_FILL
ws.column_dimensions["Z"].width = 20
for row in range(2, N_ROWS + 1):
    ws.cell(row=row, column=26, value=f'=IF(AND(A{row}<>"Y",B{row}<>"",K{row}="C"),B{row},"")')

# ---------------------------------------------------------------
# SHEET 3: Dashboard
# ---------------------------------------------------------------
wsd = wb.create_sheet("Dashboard")
wsd.sheet_view.showGridLines = False
wsd.column_dimensions["A"].width = 62
wsd.column_dimensions["B"].width = 12
wsd.column_dimensions["C"].width = 46

def section_title(row, text):
    c = wsd.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
    return row + 1

def metric(row, label, formula, note=""):
    c1 = wsd.cell(row=row, column=1, value=label)
    c1.font = Font(name=FONT_NAME, size=10)
    c2 = wsd.cell(row=row, column=2, value=formula)
    c2.font = Font(name=FONT_NAME, size=10, bold=True)
    c2.alignment = Alignment(horizontal="center")
    c2.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    if note:
        c3 = wsd.cell(row=row, column=3, value=note)
        c3.font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
        c3.alignment = WRAP
    return row + 1

R = "$2:$" + str(N_ROWS)  # not used directly; ranges built per-column below

def rng(col):
    return f"Data!{col}2:{col}{N_ROWS}"

row = 1
c = wsd.cell(row=row, column=1, value="APS + Cancer Anticoagulation — Feasibility Dashboard")
c.font = Font(name=FONT_NAME, bold=True, size=14)
row += 1
c = wsd.cell(row=row, column=1, value="Recalculates automatically from the 'Data' tab. Worked-example rows are excluded.")
c.font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
row += 2

row = section_title(row, "1. Overall volume of evidence")
row = metric(row, "Total patient-rows entered (excluding examples)",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("B")},"<>")',
             "Every case-report/series patient and every cohort/RCT patient counted once each.")
row = metric(row, "Distinct studies/articles entered",
             f'=SUMPRODUCT((({rng("Y")})<>"")/COUNTIF({rng("Y")},{rng("Y")}&""))',
             "Counts unique Study_ID values — the number that matters for 'how many independent sources do we have', not the raw patient count.")
row += 1

row = section_title(row, "2. Split by population group (Section 1.4 of the manuscript)")
row = metric(row, "Group A — oncology patients, simple/transient aPL positivity",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"A")')
row = metric(row, "Group B — oncology patients, clinically defined APS",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"B")')
groupC_row = row
row = metric(row, "Group C — oncology patients, high-risk/triple-positive APS",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"C")',
             "This is the population the review is actually about — everything below focuses on it.")
distinctC_row = row
row = metric(row, "Distinct studies contributing at least one Group C patient",
             f'=SUMPRODUCT((({rng("Z")})<>"")/COUNTIF({rng("Z")},{rng("Z")}&""))')
row += 1

row = section_title(row, "3. Group C — study design mix")
row = metric(row, "Group C patients from case reports or case series",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("D")},"Case report")'
             f'+COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("D")},"Case series")')
comparativeDesign_row = row
row = metric(row, "Group C patients from cohorts, registries, or RCTs",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("D")},"Retrospective cohort")'
             f'+COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("D")},"Prospective cohort")'
             f'+COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("D")},"Registry")'
             f'+COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("D")},"RCT")',
             "A non-zero count here is what would allow a genuine comparative effect estimate, not just a descriptive proportion.")
row += 1

row = section_title(row, "4. Group C — anticoagulant strategy (comparator check)")
vka_row = row
row = metric(row, "Group C patients treated with VKA",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"VKA")')
doac_row = row
row = metric(row, "Group C patients treated with a DOAC",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"DOAC")')
row = metric(row, "Group C patients treated with LMWH",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"LMWH")')
row += 1

row = section_title(row, "5. Group C — outcomes by strategy")
row = metric(row, "VKA: recurrence Y / bleeding (major or CRNMB) Y",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"VKA",{rng("R")},"Y")&" / "&'
             f'(COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"VKA",{rng("S")},"Y")'
             f'+COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"VKA",{rng("T")},"Y"))')
row = metric(row, "DOAC: recurrence Y / bleeding (major or CRNMB) Y",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"DOAC",{rng("R")},"Y")&" / "&'
             f'(COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"DOAC",{rng("S")},"Y")'
             f'+COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"DOAC",{rng("T")},"Y"))')
row = metric(row, "LMWH: recurrence Y / bleeding (major or CRNMB) Y",
             f'=COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"LMWH",{rng("R")},"Y")&" / "&'
             f'(COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"LMWH",{rng("S")},"Y")'
             f'+COUNTIFS({rng("A")},"<>Y",{rng("K")},"C",{rng("P")},"LMWH",{rng("T")},"Y"))')
row += 1

row = section_title(row, "6. Rule-of-thumb feasibility gates (heuristics, not statistical proof — see Legend)")
row = metric(row, "≥5 distinct studies with a Group C patient?",
             f'=IF(B{distinctC_row}>=5,"Reached — worth exploring descriptive pooling","Not yet — keep collecting")',
             f"References B{distinctC_row} ('Distinct studies contributing at least one Group C patient').")
row = metric(row, "≥10 total Group C patients?",
             f'=IF(B{groupC_row}>=10,"Reached","Not yet — keep collecting")',
             f"References B{groupC_row} ('Group C' count in Section 2).")
row = metric(row, "≥3 Group C patients on BOTH VKA and DOAC?",
             f'=IF(AND(B{vka_row}>=3,B{doac_row}>=3),"Yes — a descriptive strategy comparison becomes defensible","No — cannot compare strategies yet")',
             f"References B{vka_row}/B{doac_row} (VKA / DOAC counts in Section 4).")
row = metric(row, "Any genuine comparative design (cohort/registry/RCT) in Group C?",
             f'=IF(B{comparativeDesign_row}>0,"Yes — a true comparative meta-analysis may be possible","No — only case-level descriptive synthesis is honest, whatever the case count")',
             f"References B{comparativeDesign_row} ('cohorts, registries, or RCTs' count in Section 3).")
row += 1

row = section_title(row, "7. Bottom line")
c1 = wsd.cell(row=row, column=1,
              value="Even when every gate above is 'Reached', case report/series volume alone supports a descriptive "
                    "systematic review with a pooled proportion at most — never a pooled comparative effect size — "
                    "because case reports have no denominator, no control group, and strong publication bias "
                    "(unusual or dramatic courses are reported far more often than uneventful ones). A genuine "
                    "meta-analysis of DOAC vs VKA (or vs LMWH) effect on recurrence/bleeding in Group C requires "
                    "Section 3's 'cohorts, registries, or RCTs' count to be non-zero, not just a high case count.")
c1.font = Font(name=FONT_NAME, size=10, italic=True)
c1.alignment = WRAP
wsd.row_dimensions[row].height = 60
wsd.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

wb.save("/home/user/research-uniGE/manuscript/APS_Cancer_Literature_Tracker.xlsx")
print("done")
